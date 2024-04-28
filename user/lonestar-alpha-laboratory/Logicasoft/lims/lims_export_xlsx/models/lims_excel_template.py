# -*- coding: utf-8 -*-

##############################################################################
#
#    Odoo Proprietary License v1.0
#
#    Copyright (c) 2013 LogicaSoft SPRL (<http://www.logicasoft.eu>).
#
#    This software and associated files (the "Software") may only be used (executed,
#    modified, executed after modifications) if you have purchased a valid license
#    from the authors, typically via Odoo Apps, or if you have received a written
#    agreement from the authors of the Software.
#
#    You may develop Odoo modules that use the Software as a library (typically
#    by depending on it, importing it and using its resources), but without copying
#    any source code or material from the Software. You may distribute those
#    modules under the license of your choice, provided that this license is
#    compatible with the terms of the Odoo Proprietary License (For example:
#    LGPL, MIT, or proprietary licenses similar to this one).
#
#    It is forbidden to publish, distribute, sublicense, or sell copies of the Software
#    or modified copies of the Software.
#
#    The above copyright notice and this permission notice must be included in all
#    copies or substantial portions of the Software.
#
#    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
#    IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
#    DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
#    ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
#    DEALINGS IN THE SOFTWARE.
#
##############################################################################
from odoo import fields, models, exceptions, _, api
import os
import base64
import tempfile
from odoo.tools.safe_eval import safe_eval
from openpyxl import load_workbook
from openpyxl.comments import Comment


class LimsExcelTemplate(models.Model):
    _name = 'lims.excel.template'
    _description = 'Lims excel template'
    _order = 'id desc'
    _rec_name = 'rel_attachment_name'

    @api.model
    def get_default_code(self):
        return """
cells = []
if record._name == 'lims.sop' or record._name  == 'lims.analysis':
    nu_result_ids = record.result_num_ids.filtered(lambda r:r.rel_type in ['done','validated'])
    ca_result_ids = record.result_compute_ids.filtered(lambda r:r.rel_type in ['done','validated'])
    se_result_ids = record.result_sel_ids.filtered(lambda r:r.rel_type in ['done','validated'])
    tx_result_ids = record.result_text_ids.filtered(lambda r:r.rel_type in ['done','validated'])
    
    for res in nu_result_ids:
        if res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference:
            cell = res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference
            cell_value = res.corrected_value
            cell_comment = False
            cells.append({
                'coord': cell, 
                'value': cell_value, 
                'comment': res.comment if res.show else ''
            })
    for res in ca_result_ids:
        if res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference:
            cell = res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference
            cell_value = res.value
            cells.append({
                'coord': cell, 
                'value': cell_value, 
                'comment': res.comment if res.show else ''
            })
    for res in se_result_ids:
        if res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference:
            cell = res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference
            cell_value = res.value_id.name
            cells.append({
                'coord': cell, 
                'value': cell_value, 
                'comment': res.comment if res.show else ''
            })
    for res in tx_result_ids:
        if res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference:
            cell = res.method_param_charac_id.export_reference or res.method_param_charac_id.parameter_id.export_reference
            cell_value = res.value
            cells.append({
                'coord': cell, 
                'value': cell_value, 
                'comment': res.comment if res.show else ''
            })
"""

    sequence = fields.Integer(default=1)
    active = fields.Boolean('active', default=1)
    ir_attachment_id = fields.Many2one('ir.attachment', string='Related attachment',
                                       domain=['|',
                                               ('mimetype', '=',
                                                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                                               ('mimetype', '=', 'application/vnd.ms-excel.sheet.macroEnabled.12')],
                                       help='Use only file with extension .xlsx')
    rel_attachment_name = fields.Char(related='ir_attachment_id.name', store=True)
    code = fields.Text('Code', default=get_default_code)

    def fill_excel_template(self, context=False):
        attachment_obj = self.env['ir.attachment']
        at_in = self.ir_attachment_id
        if at_in and at_in.mimetype not in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                            'application/vnd.ms-excel.sheet.macroEnabled.12']:
            raise exceptions.UserError(_("Wrong format ! Excel file must be '.xlsx'"))
        at_ext = '.xlsx'
        if at_in.mimetype == 'application/vnd.ms-excel.sheet.macroEnabled.12':
            at_ext = '.xlsm'
        if context:
            res_model = context.get('active_model')
            res_id = context.get('active_id')
            doc_name = context.get('doc_name')
            store_fname = f"{doc_name}_{at_in.name}"
        else:
            res_model = False
            res_id = False
            store_fname = f"{at_in.name}{at_ext}"
            doc_name = store_fname

        try:
            temp_dir = tempfile.TemporaryDirectory()
            f_xl_in_path = os.path.join(temp_dir.name, f"f_xl_in{at_ext}")
            with open(f_xl_in_path, 'wb') as f_xl_in:
                f_xl_in.write(base64.decodebytes(at_in.datas))
            if at_ext == '.xlsm':
                wb = load_workbook(f_xl_in_path, keep_vba=True)
            else:
                wb = load_workbook(f_xl_in_path)
            sht0 = wb[wb.sheetnames[0]]
            code = self.code or self.get_default_code()
            if res_model and res_id:
                record = self.env[res_model].browse(res_id)
                ctx = {'record': record}
                safe_eval(code, ctx, mode="exec", nocopy=True)
                if ctx and ctx.get('cells'):
                    cells = ctx.get('cells')
                    for cell in cells:
                        if cell.get('coord') and cell.get('value'):
                            sht0[cell.get('coord')].value = cell.get('value')
                            if cell.get('comment'):
                                sht0[cell.get('coord')].comment = Comment(cell.get('comment'), 'Odoo LIMS')
            f_xl_out_path = os.path.join(temp_dir.name, f"f_xl_out{at_ext}")
            wb.save(f_xl_out_path)
            wb.close()

            with open(f_xl_out_path, 'rb') as f_xl_out:
                binary_content = f_xl_out.read()
                b64_content = base64.encodebytes(binary_content)
            return attachment_obj.create({
                'res_model': res_model or False,
                'res_id': res_id,
                'name': f"{doc_name}{at_ext}",
                'datas': b64_content,
                'store_fname': store_fname
            })

        except Exception as e:
            raise exceptions.UserError(_("Something went wrong when : \n{}").format(e)) from e

        finally:
            temp_dir.cleanup()
