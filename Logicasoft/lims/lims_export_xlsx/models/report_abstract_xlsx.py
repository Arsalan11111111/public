# -*- coding: utf-8 -*-
import base64

##############################################################################
#
#    Odoo Proprietary License v1.0
#
#    Copyright (c) 2013 LogicaSoft SPRL (<http://www.logicasoft.eu>).
#
#    This software and associated files (the "Software") may only be used (executed,
#    modified, executed after modifications) if you have purchased a valid license
#    from the authors, typically via Odoo Apps, or if you have received a written
#    agreement from the authors of the Software.
#
#    You may develop Odoo modules that use the Software as a library (typically
#    by depending on it, importing it and using its resources), but without copying
#    any source code or material from the Software. You may distribute those
#    modules under the license of your choice, provided that this license is
#    compatible with the terms of the Odoo Proprietary License (For example:
#    LGPL, MIT, or proprietary licenses similar to this one).
#
#    It is forbidden to publish, distribute, sublicense, or sell copies of the Software
#    or modified copies of the Software.
#
#    The above copyright notice and this permission notice must be included in all
#    copies or substantial portions of the Software.
#
#    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
#    IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
#    DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
#    ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
#    DEALINGS IN THE SOFTWARE.
#
##############################################################################
from odoo import models

import logging
from io import BytesIO

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import load_workbook

except ImportError:
    _logger.debug("Can not import xlsxwriter`.")


class ReportXlsxAbstract(models.AbstractModel):
    _inherit = "report.report_xlsx.abstract"
    _description = "Abstract XLSX Report"

    def create_xlsx_report(self, docids, data, library=None, template_file=None):
        if library != 'openpyxl':
            return super().create_xlsx_report(docids, data)
        file_ext = "xlsx"
        objs = self._get_objs_for_report(docids, data)
        file_data = BytesIO()
        file_ext, workbook = self._get_excel_workbook(file_ext, template_file)
        self.generate_xlsx_report(workbook, data, objs, template_file=template_file)
        workbook.save(file_data)
        file_data.seek(0)
        del workbook
        return file_data.read(), file_ext

    def _get_excel_workbook(self, file_ext, template_file):
        if template_file and template_file.datas and template_file.mimetype in [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel.sheet.macroEnabled.12']:
            keep_vba = False
            if template_file.mimetype == 'application/vnd.ms-excel.sheet.macroEnabled.12':
                file_ext = '.xlsm'
                keep_vba = True
            xlsx = BytesIO()
            xlsx.write(base64.decodebytes(template_file.datas))
            workbook = load_workbook(filename=xlsx, keep_vba=keep_vba)
        else:
            workbook = openpyxl.Workbook()
        return file_ext, workbook

    def generate_xlsx_report(self, workbook, data, objs, template_file=None):
        raise NotImplementedError()
