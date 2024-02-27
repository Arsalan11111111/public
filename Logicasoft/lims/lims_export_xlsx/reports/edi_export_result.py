# -*- coding: utf-8 -*-
##############################################################################
#
#    Odoo Proprietary License v1.0
#
#    Copyright (c) 2013 LogicaSoft SPRL (<http://www.logicasoft.eu>).
#
#    This software and associated files (the "Software") may only be used (executed,
#    modified, executed after modifications) if you have purchased a valid license
#    from the authors, typically via Odoo Apps, or if you have received a written
#    agreement from the authors of the Software.
#
#    You may develop Odoo modules that use the Software as a library (typically
#    by depending on it, importing it and using its resources), but without copying
#    any source code or material from the Software. You may distribute those
#    modules under the license of your choice, provided that this license is
#    compatible with the terms of the Odoo Proprietary License (For example:
#    LGPL, MIT, or proprietary licenses similar to this one).
#
#    It is forbidden to publish, distribute, sublicense, or sell copies of the Software
#    or modified copies of the Software.
#
#    The above copyright notice and this permission notice must be included in all
#    copies or substantial portions of the Software.
#
#    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
#    IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
#    DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
#    ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
#    DEALINGS IN THE SOFTWARE.
#
##############################################################################
from odoo import models
from collections import OrderedDict

from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class EdiExportResult(models.AbstractModel):
    _name = 'report.lims_report_xls.report_edi_export_result'
    _description = 'EDI export result report'
    _inherit = 'report.report_xlsx.abstract'

    def generate_xlsx_report(self, workbook, data, element_ids, template_file=None):
        self = self.sudo()
        # main objectif found sop to treat
        sop_ids = self.env['lims.sop']
        sop_ids = self._get_sop_ids(element_ids, sop_ids)
        result_ids = self._get_result_ids(sop_ids)
        result_ids = self._sort_result_ids(result_ids)
        current_element_id = False
        current_sheet = self._set_uom_ids(workbook, create_sheet=not bool(template_file))
        row = 1
        for result in result_ids:
            if current_element_id != result.analysis_id:
                current_element_id = result.analysis_id
                current_sheet = self._add_worksheet(current_element_id, workbook) or self._get_default_worksheet(
                    workbook, template_file)
                current_sheet.protection.formatColumns = False
                current_sheet.protection.formatRows = False
                current_sheet.protection.autoFilter = False
                current_sheet.protection.enable()
                row = 1

            datas = result.get_export_values(ordered_dictionary=self._get_header_list())
            for col, name in enumerate(datas):
                current_data = datas.get(name)
                current_cell = current_sheet[f'{get_column_letter(col + 1)}{row + 1}']
                current_cell.value = current_data.get('value')
                cell_validation = current_data.get('validate')
                cell_protection = current_data.get('protection')
                if cell_validation:
                    self.set_validation_cell(current_cell, current_sheet, cell_validation.copy())
                if cell_protection == 'unlock':
                    current_cell.protection = Protection(locked=False)
            row += 1

    def _set_uom_ids(self, workbook, create_sheet=False):
        current_sheet = self._get_default_worksheet(workbook, create=create_sheet)
        uom_ids = self.env['uom.uom'].search([])
        uom_names = [uom_id.name for uom_id in uom_ids]
        uom_names.insert(0, '_uom_names')
        self._set_defined_list_in_odoo_worksheet(uom_names, workbook)
        current_sheet.protection.enable()
        current_sheet.sheet_state = 'veryHidden'
        return current_sheet

    def set_validation_cell(self, cell, current_sheet, dictionary):
        # Force data validation for a number in this cell.
        dictionary = dictionary.copy()
        if dictionary.get('type') == 'custom' and '{}' in dictionary.get('formula1'):
            dictionary['formula1'] = dictionary['formula1'].format(cell.coordinate)
        elif dictionary.get('type') == 'list' and dictionary.get('defined_name') and dictionary.get('values'):
            defined_name = dictionary.pop('defined_name')
            values = dictionary.pop('values')
            current_sheet.parent.defined_names.get(defined_name)
            if not current_sheet.parent.defined_names.get(defined_name):
                values.insert(0, defined_name)
                self._set_defined_list_in_odoo_worksheet(values, current_sheet.parent)
        dv = DataValidation(**dictionary)
        current_sheet.add_data_validation(dv)
        dv.add(cell)

    def _set_defined_list_in_odoo_worksheet(self, list_element, workbook):
        header = list_element[0] if list_element else False
        self._set_defined_name_in_odoo_worksheet(header, list_element, workbook)
        for element in list_element:
            if not isinstance(element, list):
                element = [element]
            workbook['_odoo_'].append(element)

    def _set_defined_name_in_odoo_worksheet(self, header, list_element, workbook, sheet='_odoo_'):
        row_start = workbook[sheet].max_row
        len_elements = len(list_element)
        row_start = 1 if row_start == 1 else row_start + 1
        row_end = row_start + len_elements - 1
        workbook.create_named_range(header, workbook[sheet],
                                    f'${get_column_letter(1)}${row_start + 1}:${get_column_letter(1)}${row_end}')

    def _get_default_worksheet(self, workbook, sheet='_odoo_', create=False):
        if create:
            sheet = workbook.worksheets[0]
            sheet.name = '_odoo_'
            sheet.title = '_odoo_'
        return workbook['_odoo_'] if '_odoo_' in workbook.sheetnames else workbook.create_sheet('_odoo_')

    def _add_worksheet(self, current_element_id, workbook):
        current_sheet = False
        if current_element_id:
            current_sheet = workbook.create_sheet(current_element_id.name)
            current_sheet = self._set_worksheet_header(current_sheet)
        return current_sheet

    def _set_worksheet_header(self, current_sheet):
        current_sheet.freeze_panes = 'A2'
        ordered_dict = self._get_header_list()
        current_sheet.append([header.get('label') for header in ordered_dict.values()])
        return current_sheet

    def _get_header_list(self):
        return OrderedDict(
            {'result': {'label': 'Result'},
             'request': {'label': 'Request'},
             'analysis': {'label': 'Analysis'},
             'test': {'label': 'Test'},
             'parameter': {'label': 'Parameter'},
             'value': {'label': 'Value'},
             'uom': {'label': 'UoM'},
             'state': {'label': 'State'},
             'stage': {'label': 'Stage'},
             'comment': {'label': 'Comment'},
             'dilution': {'label': 'Dilution'},
             'lod': {'label': 'LOD'},
             'loq': {'label': 'LOQ'},
             'mloq': {'label': 'mLOQ'},
             }
        )

    def _sort_result_ids(self, result_ids):
        result_ids = sorted(result_ids,
                            key=lambda r: (r.analysis_id.name, r.sop_id.name, r.method_param_charac_id.tech_name))
        return result_ids

    def _get_result_ids(self, sop_ids):
        return sop_ids.get_results_filtered(domain=lambda r: r.rel_type not in ['cancel', 'rework'])

    def _get_sop_ids(self, element_ids, sop_ids):
        if element_ids._name == 'lims.analysis.request':
            sop_ids = element_ids.get_analyses_filtered(remove_stage=['cancel']).get_sops_filtered(
                remove_stage=['cancel'])
        elif element_ids._name == 'lims.analysis':
            sop_ids = element_ids.get_sops_filtered(remove_stage=['cancel'])
        elif element_ids._name == 'lims.sop':
            sop_ids = element_ids or self.env['lims.sop']
        return sop_ids


class EdiExportResultSop(models.AbstractModel):
    _name = 'report.lims_report_xls.report_edi_export_result_sop'
    _inherit = 'report.lims_report_xls.report_edi_export_result'
    _description = 'report_edi_export_result_sop'


class EdiExportResultAnalysis(models.AbstractModel):
    _name = 'report.lims_report_xls.report_edi_export_result_analysis'
    _inherit = 'report.lims_report_xls.report_edi_export_result'
    _description = 'report_edi_export_result_analysis'


class EdiExportResultRequest(models.AbstractModel):
    _name = 'report.lims_report_xls.report_edi_export_result_request'
    _inherit = 'report.lims_report_xls.report_edi_export_result'
    _description = 'report_edi_export_result_request'
